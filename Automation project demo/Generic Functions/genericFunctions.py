from datetime import datetime
import xlrd
import logging
import openpyxl

class generic(object):
    
    wb=None
    destination=None
    
    exePath=None
    '''def __init__(self,dest=None):
        self.dest=dest
        
        wb = openpyxl.load_workbook(self.dest) '''
    
    
    r'''
    def getdestination(cls,dest):
        generic.destination=dest
        generic.wb= openpyxl.load_workbook(generic.destination)
        print ("generic.wb: ", generic.wb)
    '''
    @classmethod
    def getdestination(cls,exePath):
        generic.destination=exePath+"viaData_stage_execution.xlsx"
        
        generic.wb= openpyxl.load_workbook(generic.destination)
        print ("generic.wb: ", generic.wb)    
        
    #Mmethod to get current time
    @staticmethod
    def getCurrentTime():
        time=datetime.time(datetime.now()).strftime('%H-%M-%S')
        return time
    
    #Method to get current time    
    @staticmethod   
    def test_getCurrentDate():
        date=datetime.date(datetime.now()).strftime('%d/%m/%y')
        return date  
    
    #Method to check execution flag in the excel. A test case will be executed only when execution flag is "Yes"
    @staticmethod
    def checkExecutionFlag(sheet0, TCName):
        exeFlag=generic.fetchValueFromExcel(sheet0, TCName, "ExecutionFlag")
        'print ("exeFlag is: ", exeFlag)'
        return exeFlag
            
    #Method to fetch value from the excel    
    @staticmethod
    def fetchValueFromExcel(sheet1, TCName, reqVal):
        wb=xlrd.open_workbook(".\\Test_data\\viadata_stage.xlsx")
        sheet=wb.sheet_by_name(sheet1)
        row_count=sheet.nrows
        column_count=sheet.ncols
        for i in range(row_count):
            for j in range(column_count):
                cell=sheet.cell(i,j).value
                if(cell==TCName):
                    requiredRow=i
                    break
        
        for k in range(column_count):
            cell1=sheet.cell(0,k).value
            if(cell1==reqVal):
                requiredCol=k
                break
        'print ("sheet.cell(i,k).value is: ", sheet.cell(i,k).value) '   
        return sheet.cell(i,k).value        
    
    #Log creation
    @staticmethod                
    def logCreation():
        
        # create logger
        logger = logging.getLogger('viaApplicationLogger')
        logger.setLevel(logging.DEBUG)

        # create console handler and set level to debug
        ch = logging.StreamHandler()
        ch.setLevel(logging.DEBUG)
        
        fh=logging.FileHandler(".\\Logs_via.log")
        fh.setLevel(logging.DEBUG)

        # create formatter
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

        # add formatter to ch
        ch.setFormatter(formatter)
        
        fh.setFormatter(formatter)

        # add ch to logger
        logger.addHandler(ch)
        
        logger.addHandler(fh)

        # 'application' code
        r'''
        logger.debug('debug message')
        logger.info('info message')
        logger.warn('warn message')
        logger.error('error message')
        logger.critical('critical message')
        '''                        
        return logger                      
    #Method to capture screenshot                           
    @staticmethod
    def captureScreenshot(driver, TCName ):
        driver.save_screenshot(".\\Result\\"+TCName+".png")
        

    #Method to set value to excel    
    @staticmethod
    def setValueToExcel(sheet0, TCName , columnName, cellValue): 
        print ("setValue in excel: ", cellValue)   
        sheet = generic.wb.get_sheet_by_name(sheet0)
        row_count=sheet.max_row
        col_count=sheet.max_column
        for i in range(1,row_count+1):
            for j in range(1,col_count+1):
                if(sheet.cell(i, j).value==TCName):
                    requiredRow=i
                    print("requiredRow is: ", requiredRow)
                    break
                
            for k in range(1,col_count+1):
                if(sheet.cell(1,k).value==columnName):
                    requiredCol=k
                    print ("requiredCol is: ", requiredCol)
                    break 
        cell = sheet.cell(requiredRow, requiredCol)
            
        cell.value = cellValue
        print ("cell.value is: ", cell.value)
        generic.wb.save(generic.destination)  
